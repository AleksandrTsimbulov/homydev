from openpyxl import load_workbook
from collections import namedtuple


class ExcelExtractor:
    def __init__(self, excel_translems_file, min_row=5, max_col=2, max_row=10000, topic_cell='B3'):
        self._min_row = min_row
        self._max_col = max_col
        self._max_row = max_row
        self._topic_cell = topic_cell
        self._working_sheet = load_workbook(filename=f'{excel_translems_file}').active

    def get_translems(self):
        translem_list = []
        Translem = namedtuple('Translem', 'english russian')
        for excel_row in self._working_sheet.iter_rows(min_row=self._min_row,
                                                       max_col=self._max_col,
                                                       max_row=self._max_row):
            if not excel_row[0].value:
                break
            new_translem = Translem(excel_row[0].value, excel_row[1].value)
            translem_list.append((new_translem.english, new_translem.russian))
        no_repeats_translem_list = list(set(translem_list))
        return no_repeats_translem_list

    def get_topic(self):
        return self._working_sheet[f'{self._topic_cell}'].value


if __name__ == '__main__':
    parser = ExcelExtractor('Translems.xlsx')
    print(parser.get_translems())
    print(parser.get_topic())
